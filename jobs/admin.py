from django.contrib import admin
from django.http import HttpResponse
from .models import Candidate
import csv
import datetime

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'age', 'skills', 'college_name', 'passing_year', 'submitted_at')
    search_fields = ('name', 'email', 'skills', 'college_name')
    list_filter = ('passing_year',)
    readonly_fields = ('submitted_at',)
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        """Admin action to export selected candidates into an Excel file (xlsx). Falls back to CSV if openpyxl not installed."""
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Candidates'
            headers = ['Name', 'Email', 'Age', 'Skills', 'GitHub', 'Address', 'College Name', 'Passing Year', 'Phone', 'Resume URL', 'Submitted At']
            ws.append(headers)
            for cand in queryset:
                row = [
                    cand.name,
                    cand.email,
                    cand.age,
                    cand.skills,
                    cand.github_link,
                    cand.address,
                    cand.college_name,
                    cand.passing_year,
                    cand.phone_number,
                    request.build_absolute_uri(cand.resume.url) if cand.resume else '',
                    cand.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                ]
                ws.append(row)
            # Autosize columns
            for i, _ in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].auto_size = True
            now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'candidates_{now}.xlsx'
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            wb.save(response)
            return response
        else:
            # Fallback to CSV
            now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'candidates_{now}.csv'
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            writer = csv.writer(response)
            writer.writerow(['Name', 'Email', 'Age', 'Skills', 'GitHub', 'Address', 'College Name', 'Passing Year', 'Phone', 'Resume URL', 'Submitted At'])
            for cand in queryset:
                writer.writerow([
                    cand.name,
                    cand.email,
                    cand.age,
                    cand.skills,
                    cand.github_link,
                    cand.address,
                    cand.college_name,
                    cand.passing_year,
                    cand.phone_number,
                    request.build_absolute_uri(cand.resume.url) if cand.resume else '',
                    cand.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                ])
            return response

    export_as_excel.short_description = 'Export selected candidates to Excel/CSV'
